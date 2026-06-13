"""
Módulo de exportación de facturas a CSV y Excel.
"""
import csv
from pathlib import Path
from typing import List, Dict, Any
from datetime import datetime
import logging

logger = logging.getLogger("FacturasGanaTodo.export")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl no disponible - exportación a Excel deshabilitada")


def export_to_csv(facturas: List[Dict[str, Any]], filepath: str) -> bool:
    """
    Exporta facturas a archivo CSV.
    
    Args:
        facturas: Lista de diccionarios con datos de facturas
        filepath: Ruta del archivo CSV a crear
        
    Returns:
        True si la exportación fue exitosa
    """
    try:
        # Definir columnas
        fieldnames = [
            'id', 'numero_factura', 'proveedor', 'valor', 
            'fecha_vencimiento', 'estado', 'notas',
            'hora_alerta_1', 'hora_alerta_2', 'hora_alerta_3'
        ]
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            
            for factura in facturas:
                # Filtrar solo las columnas que queremos exportar
                row = {k: factura.get(k, '') for k in fieldnames}
                writer.writerow(row)
        
        logger.info(f"✓ Exportadas {len(facturas)} facturas a CSV: {filepath}")
        return True
    
    except Exception as e:
        logger.exception(f"Error al exportar a CSV: {e}")
        return False


def export_to_excel(facturas: List[Dict[str, Any]], filepath: str) -> bool:
    """
    Exporta facturas a archivo Excel con formato profesional.
    
    Args:
        facturas: Lista de diccionarios con datos de facturas
        filepath: Ruta del archivo Excel a crear
        
    Returns:
        True si la exportación fue exitosa
    """
    if not EXCEL_AVAILABLE:
        logger.error("openpyxl no disponible")
        return False
    
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Facturas"
        
        # ORDENAR por ID (orden de ingreso) - menor ID primero
        facturas_ordenadas = sorted(facturas, key=lambda x: x.get('id', 0))
        
        # Definir headers con nuevas columnas
        headers = [
            'ID', 'Número Factura', 'Proveedor', 'Valor', 
            'Fecha Vencimiento', 'Estado', 'Vencida', 'Tiene Comprobante', 'Tiene PDF', 'Notas',
            'Alerta 1', 'Alerta 2', 'Alerta 3'
        ]
        
        # Escribir headers con formato
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Escribir datos
        for row_idx, factura in enumerate(facturas_ordenadas, start=2):
            # ID
            cell = ws.cell(row=row_idx, column=1)
            cell.value = factura.get('id', '')
            cell.alignment = Alignment(horizontal='center')
            
            # Número Factura
            cell = ws.cell(row=row_idx, column=2)
            cell.value = factura.get('numero_factura', '')
            cell.font = Font(bold=True)
            
            # Proveedor
            cell = ws.cell(row=row_idx, column=3)
            cell.value = factura.get('proveedor', '')
            
            # Valor con formato de moneda
            cell = ws.cell(row=row_idx, column=4)
            valor = factura.get('valor', 0)
            cell.value = float(valor) if valor else 0
            cell.number_format = '$#,##0.00'
            cell.alignment = Alignment(horizontal='right')
            
            # Fecha de vencimiento
            cell = ws.cell(row=row_idx, column=5)
            fecha_str = factura.get('fecha_vencimiento', '')
            if fecha_str:
                try:
                    # Parsear y formatear fecha
                    dt = datetime.fromisoformat(fecha_str)
                    cell.value = dt.strftime('%d/%m/%Y %I:%M %p')
                except:
                    cell.value = fecha_str
            else:
                cell.value = ''
            
            # Estado
            cell = ws.cell(row=row_idx, column=6)
            estado = factura.get('estado', '')
            cell.value = estado
            
            # Colorear según estado
            if estado == 'Pagada':
                cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                cell.font = Font(color="065F46", bold=True)
            elif estado == 'Pendiente':
                cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                cell.font = Font(color="92400E", bold=True)
            
            # NUEVA COLUMNA 7: Vencida
            cell = ws.cell(row=row_idx, column=7)
            is_vencida = "No"
            if fecha_str and estado == 'Pendiente':
                try:
                    dt = datetime.fromisoformat(fecha_str)
                    if dt < datetime.now():
                        is_vencida = "Sí"
                except:
                    pass
            cell.value = is_vencida
            cell.alignment = Alignment(horizontal='center')
            if is_vencida == "Sí":
                cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                cell.font = Font(color="991B1B", bold=True)
            
            # NUEVA COLUMNA 8: Tiene Comprobante
            cell = ws.cell(row=row_idx, column=8)
            # Si está pagada, es probable que tenga comprobante
            tiene_comprobante = "Sí" if estado == 'Pagada' else "No"
            cell.value = tiene_comprobante
            cell.alignment = Alignment(horizontal='center')
            
            # NUEVA COLUMNA 9: Tiene PDF
            cell = ws.cell(row=row_idx, column=9)
            tiene_pdf = "Sí" if factura.get('pdf_path') else "No"
            cell.value = tiene_pdf
            cell.alignment = Alignment(horizontal='center')
            if tiene_pdf == "Sí":
                cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                cell.font = Font(color="065F46", bold=True)
            
            # Notas (columna 10)
            cell = ws.cell(row=row_idx, column=10)
            notas = factura.get('notas', '')
            cell.value = notas[:100] if notas else ''
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Alertas (columnas 11, 12, 13)
            cell = ws.cell(row=row_idx, column=11)
            cell.value = factura.get('hora_alerta_1', '')
            cell.alignment = Alignment(horizontal='center')
            
            cell = ws.cell(row=row_idx, column=12)
            cell.value = factura.get('hora_alerta_2', '')
            cell.alignment = Alignment(horizontal='center')
            
            cell = ws.cell(row=row_idx, column=13)
            cell.value = factura.get('hora_alerta_3', '')
            cell.alignment = Alignment(horizontal='center')
        
        # Ajustar ancho de columnas
        column_widths = {
            'A': 8,   # ID
            'B': 18,  # Número Factura
            'C': 25,  # Proveedor
            'D': 15,  # Valor
            'E': 20,  # Fecha
            'F': 12,  # Estado
            'G': 10,  # Vencida
            'H': 18,  # Tiene Comprobante
            'I': 12,  # Tiene PDF
            'J': 40,  # Notas
            'K': 12,  # Alerta 1
            'L': 12,  # Alerta 2
            'M': 12,  # Alerta 3
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Altura de la fila de header
        ws.row_dimensions[1].height = 25
        
        # Congelar primera fila
        ws.freeze_panes = 'A2'
        
        # Guardar
        wb.save(filepath)
        
        logger.info(f"✓ Exportadas {len(facturas)} facturas a Excel: {filepath}")
        return True
    
    except Exception as e:
        logger.exception(f"Error al exportar a Excel: {e}")
        return False


def get_export_filename(extension: str = 'csv') -> str:
    """
    Genera nombre de archivo para exportación.
    
    Args:
        extension: Extensión del archivo ('csv' o 'xlsx')
        
    Returns:
        Nombre de archivo sugerido
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"facturas_export_{timestamp}.{extension}"
